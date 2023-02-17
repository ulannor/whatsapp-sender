import pandas as pd
import pywhatkit as pw
import time
from openpyxl import load_workbook
import keyboard as k
import random


def format_phone(phone):
    if len(phone) == 10 and phone[0] == '0':
        phone = phone.replace('0', '996', 1)
        return phone
    elif len(phone) == 9:
        phone = phone + '996'
        return phone
    for i in phone:
        if i in '-() +;':
            phone = phone.replace(i, '')
            return '+' + phone
        else:
            return phone
    phone = '+' + phone.replace('.0', '')
    return phone


def read_txt_file(txtpath):
    with open(txtpath, 'r', encoding='utf-8') as f:
        text_msg = ''
        for row in f:
            text_msg += row
    return text_msg


def send_msg(phone, msg):
    sec = random.randint(40, 60)
    phone = format_phone(phone)
    pw.sendwhatmsg_instantly(phone, msg, wait_time=sec, close_time=random.randint(5, 10))
    time.sleep(random.randint(10, 15))
    k.press_and_release('ctrl+w')

def logwriter(dftemp, logpath):
    workbook = load_workbook(logpath)
    writer = pd.ExcelWriter(logpath, engine='openpyxl')
    writer.book = workbook
    worksheet = workbook.active
    dftemp = dftemp.transpose()
    dftemp.to_excel(writer, startrow=worksheet.max_row, startcol=0, index = False, header= False)
    writer.close()