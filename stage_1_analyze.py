import pandas as pd
from collections import Counter
import openpyxl as op
from openpyxl.styles import PatternFill, Font
from openpyxl.formatting import Rule
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.utils import get_column_letter

# Set file names and paths
FILENAME = 'нотариальные услуги'
SOURCE_PATH = f'.\\sourcedata\\{FILENAME}.xlsx'
LOG_PATH = f'.\\wa_mailout_log\\wa_mailout_log.xlsx'
TEMP_PATH = f'.\\temp\\{FILENAME}_edited.xlsx'
EDITED_PATH = f'.\\edited\\{FILENAME}_edited.xlsx'

# Read source data into pandas dataframe
df = pd.read_excel(SOURCE_PATH, dtype='object')

# Read log data into pandas dataframe
dfLog = pd.read_excel(LOG_PATH, dtype='object')

# Add columns for counting WhatsApp numbers and message count
if 'wanumbers_available' not in df.columns:
    df.insert(10, 'wanumbers_available', 0)
    df.insert(11, 'message_count', 0)

# Create list of WhatsApp and non-WhatsApp column names
waNumberColList = [col for col in df if col.startswith('whatsapp')]
nonWaNumberColList = [col for col in df if col.startswith('phone')]

# Count the number of available WhatsApp numbers for each client
for col in waNumberColList:
    df.loc[df[col].notnull(), 'wanumbers_available'] += 1

# Count the number of messages sent through WhatsApp-sender to each client in df
urlList = dfLog['2GIS URL'].tolist()
urlAdjustedList = [urlList[i] for i in range(len(urlList)) if (i == 0) or urlList[i] != urlList[i - 1]]
urlCounterDict = Counter(urlAdjustedList)
df["message_count"] = df["2GIS URL"].map(urlCounterDict)

# Set index for df and dfLog
dfLog['temp_index'] = dfLog['2GIS URL']
dfLog.set_index('temp_index', inplace=True)
df['temp_index'] = df['2GIS URL']
df.set_index('temp_index', inplace=True)

# Set new informational columns in dfLog to be transferred into df
dfLog['last_message_text'] = dfLog.dropna(subset=['message_text']).groupby('2GIS URL', group_keys=False)[
    'message_text'].last()
dfLog['last_message_date'] = dfLog.dropna(subset=['date_and_time']).groupby('2GIS URL', group_keys=False)[
    'date_and_time'].last()
dfLog['message_wanumbers'] = dfLog.dropna(subset=['message_wanumber']).groupby('2GIS URL', group_keys=False)[
    'message_wanumber'].apply(lambda x: list(set(x)))
dfLog['message_nonwanumbers'] = dfLog.dropna(subset=['message_nonwanumber']).groupby('2GIS URL', group_keys=False)[
    'message_nonwanumber'].apply(lambda x: list(set(x)))
dfLog['2GISid_processed'] = dfLog.dropna(subset=['organization id']).groupby('2GIS URL', group_keys=False)[
    'organization id'].apply(lambda x: len(list(set(x))))
dfLog['2GISurl_processed'] = dfLog.dropna(subset=['2GIS URL']).groupby('2GIS URL', group_keys=False)['2GIS URL'].apply(
    lambda x: len(list(set(x))))
dfLog['wanumbers_processed'] = dfLog.dropna(subset=['message_wanumber']).groupby('2GIS URL', group_keys=False)[
    'message_wanumber'].apply(lambda x: len(list(set(x))))
dfLog['nonwanumbers_processed'] = dfLog.dropna(subset=['message_nonwanumber']).groupby('2GIS URL', group_keys=False)[
    'message_nonwanumber'].apply(lambda x: len(list(set(x))))


# Drop the duplicates and keep the last of them
dfLogNoDups = dfLog.drop_duplicates(subset='2GIS URL', keep='last')

# Create a list of column names from dfLog and select a sublist of columns from indices 15 to 23 (inclusive), where the column names are expected to be found. Selected column names are then stored in a new variable called newColList.
#! Need to modify this part to refer to names instead of numbers
newColList = list(dfLog.columns)[15:23]

# Merge dfLogNoDups with df
for name in newColList:
    if name == '2GISid_processed':
        df = df.merge(dfLogNoDups[[name, 'organization id']], on='organization id', how='left')
    else:
        df = df.merge(dfLogNoDups[[name, '2GIS URL']], on='2GIS URL', how='left')

# Reinsert new columns into specified indices
for name in newColList:
    col = df.pop(name)
    df.insert(12 + newColList.index(name), col.name, col)

# Write the contents of the modified DataFrame df to an Excel file in a subdirectory called temp, with the filename of the original file appended with _edited.xlsx.
df.to_excel(TEMP_PATH, index=False)


#Create a list of phone number column indices for conditional formatting purposes by finding the index of each column name in two separate lists.
tempLst = []
for _ in waNumberColList + nonWaNumberColList:
    tempLst.append(df.columns.tolist().index(_) + 1)

#Formatting highlights any duplicate values in each column with red text and a red fill. It freezes the first row and applies an auto-filter to the worksheet.
wb = op.load_workbook(TEMP_PATH)
ws = wb.active

red_text = Font(color="9C0006")
red_fill = PatternFill(bgColor="FFC7CE")
dxf = DifferentialStyle(font=red_text, fill=red_fill)
rule = Rule(type="duplicateValues", text="highlight", dxf=dxf)
bgn = 1
end = 100000
for _ in tempLst:
    column_letter = get_column_letter(_)
    ws.conditional_formatting.add(f'{column_letter}{bgn}:{column_letter}{end}', rule)
ws.conditional_formatting.add(f'A{bgn}:A{end}', rule)
ws.auto_filter.ref = ws.dimensions
ws.freeze_panes = 'A2'

#Write the resulting dataframe into Excel file.
wb.save(EDITED_PATH)

